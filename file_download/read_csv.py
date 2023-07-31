import pandas as pd
from openpyxl import workbook , load_workbook
from datetime import datetime

def save_to_excel():
    # print(file_name)
    # # Load the data from the CSV file and filter to get 'filtered_df'
    df2 = pd.read_csv(f'Downloads/fo28JUL2023bhav.csv')
    filtered_df = df2[df2['INSTRUMENT'].str.contains('FUT')]


    wb = load_workbook('./Copy.xlsx')
    ws = wb['Data']
    # print(ws.append(df2))
    last_row = 2
    # Define the starting row to append data (e.g., 2 to skip the header row)
    # ws.delete_rows(idx = 4 )
    # last_row = ws.max_row
    # print(last_row)
    while(last_row):
        if ws.cell(row=last_row, column=3).value == None:
            break
        else:
            last_row+=1
    print(last_row)

    starting_row = 2
    # Iterate through the rows of df2 and append to the Excel sheet
    for index, row in filtered_df.iterrows():
        # Get the row number in the Excel sheet
        target_row = starting_row + index
        ex_date = row['EXPIRY_DT']
        # print(ex_date)
        # Write each column value to the corresponding cell in the Excel sheet
        ws.cell(row=target_row, column=3, value=row['INSTRUMENT'])
        ws.cell(row=target_row, column=4, value=row['SYMBOL'])
        # Parse the date string to a datetime object
        # date_obj = datetime.strptime(ex_date, "%d-%b-%Y")
        # # year_last_two_digits = date_obj.strftime("%y")
        # exp_date = date_obj.strftime("%d-%m-%Y")
        # print(new_date_str)
        
        ws.cell(row=target_row, column=5, value=ex_date)
        ws.cell(row=target_row, column=6, value=row['STRIKE_PR'])
        ws.cell(row=target_row, column=7, value=row['OPTION_TYP'])
        ws.cell(row=target_row, column=8, value=row['OPEN'])
        ws.cell(row=target_row, column=9, value=row['HIGH'])
        ws.cell(row=target_row, column=10, value=row['LOW'])
        ws.cell(row=target_row, column=11, value=row['CLOSE'])
        ws.cell(row=target_row, column=12, value=row['SETTLE_PR'])
        ws.cell(row=target_row, column=13, value=row['CONTRACTS'])
        ws.cell(row=target_row, column=14, value=row['VAL_INLAKH'])
        ws.cell(row=target_row, column=15, value=row['OPEN_INT'])
        ws.cell(row=target_row, column=16, value=row['CHG_IN_OI'])

        time_stamp = row['TIMESTAMP']
        # date_obj = datetime.strptime(time_stamp, "%d-%-%Y")
        # # year_last_two_digits = date_obj.strftime("%y")
        # time_date = date_obj.strftime("%d-%m-%Y")
        ws.cell(row=target_row, column=17, value=time_stamp)
        
    #     #

    # print(ws)
    wb.save('./Copy.xlsx')

save_to_excel()